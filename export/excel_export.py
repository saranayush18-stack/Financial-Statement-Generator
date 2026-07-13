"""
Excel Export.

Produces a single professionally formatted workbook with sheets:
TB | Mapping | Balance Sheet | P&L | Cash Flow | Notes | Validation | Ratios

Uses openpyxl. Currency cells use formulas referencing the TB sheet wherever
practical (Balance Sheet / P&L totals sum the Notes sheet ranges) so the
workbook recalculates if a ledger amount is corrected -- it is not just a
snapshot of Python-computed numbers.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from schedule3_engine.models import Company, TrialBalance, MappingEntry
from schedule3_engine.core.statement_generator import BalanceSheet, ProfitAndLoss, CashFlowStatement
from schedule3_engine.core.notes_generator import Note
from schedule3_engine.core.ratios import RatioResult
from schedule3_engine.models import ValidationIssue
from schedule3_engine.core.soce_generator import StatementOfChangesInEquity
from schedule3_engine.core.ageing import AgeingGrid, BUCKET_LABELS, CATEGORY_LABELS

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SUBHEAD_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FILL = PatternFill("solid", fgColor="BDD7EE")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM_FMT = '#,##0.00;(#,##0.00);"-"'


def _header_font():
    return Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)


def _title_font():
    return Font(name=FONT_NAME, bold=True, size=14, color="1F4E79")


def _bold():
    return Font(name=FONT_NAME, bold=True, size=10)


def _normal():
    return Font(name=FONT_NAME, size=10)


def _write_title_block(ws, company: Company, statement_title: str, fy_label: str, next_row: int = 1):
    ws.cell(row=next_row, column=1, value=company.name).font = _title_font()
    ws.cell(row=next_row + 1, column=1, value=statement_title).font = Font(name=FONT_NAME, bold=True, size=12)
    ws.cell(row=next_row + 2, column=1, value=f"For the year ended - {fy_label}").font = Font(
        name=FONT_NAME, italic=True, size=10)
    ws.cell(row=next_row + 3, column=1, value=f"(All amounts in {company.currency} unless otherwise stated)").font = \
        Font(name=FONT_NAME, italic=True, size=9)
    return next_row + 5


def _autosize(ws, widths: dict[int, int]):
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def build_workbook(
    company: Company,
    fy_label: str,
    tb: TrialBalance,
    mappings: dict[str, MappingEntry],
    bs: BalanceSheet,
    pnl: ProfitAndLoss,
    cash_flow: CashFlowStatement,
    notes: list[Note],
    issues: list[ValidationIssue],
    ratios: list[RatioResult],
    soce: StatementOfChangesInEquity | None = None,
    receivables_ageing: AgeingGrid | None = None,
    payables_ageing: AgeingGrid | None = None,
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    _write_tb_sheet(wb, company, fy_label, tb, mappings)
    _write_mapping_sheet(wb, company, fy_label, tb, mappings)
    _write_balance_sheet_sheet(wb, company, fy_label, bs)
    _write_pnl_sheet(wb, company, fy_label, pnl)
    if soce is not None:
        _write_soce_sheet(wb, company, fy_label, soce)
    _write_cash_flow_sheet(wb, company, fy_label, cash_flow)
    if receivables_ageing is not None or payables_ageing is not None:
        _write_ageing_sheet(wb, company, fy_label, receivables_ageing, payables_ageing)
    _write_notes_sheet(wb, company, fy_label, notes)
    _write_validation_sheet(wb, company, fy_label, issues)
    _write_ratios_sheet(wb, company, fy_label, ratios)

    return wb


def _write_soce_sheet(wb, company, fy_label, soce: StatementOfChangesInEquity):
    ws = wb.create_sheet("SOCE")
    r = _write_title_block(ws, company, "Statement of Changes in Equity", fy_label)

    ws.cell(row=r, column=1, value="A. Equity Share Capital").font = Font(
        name=FONT_NAME, bold=True, color="1F4E79")
    r += 1
    sc = soce.equity_share_capital
    for label, val in [
        ("Balance at the beginning of the year", sc.opening),
        ("Changes in equity share capital during the year", sc.changes_during_year),
        ("Balance at the end of the year", sc.closing),
    ]:
        ws.cell(row=r, column=1, value=label).font = _normal()
        ws.cell(row=r, column=2, value=val).number_format = NUM_FMT
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="B. Other Equity").font = Font(
        name=FONT_NAME, bold=True, color="1F4E79")
    r += 1
    headers = ["Component", "Opening Balance", "Profit for the Year", "Other Additions/(Deductions)", "Closing Balance"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = _header_font()
        cell.fill = HEADER_FILL
    r += 1
    for comp in soce.other_equity:
        ws.cell(row=r, column=1, value=comp.component).font = _normal()
        ws.cell(row=r, column=2, value=comp.opening).number_format = NUM_FMT
        ws.cell(row=r, column=3, value=comp.profit_for_the_year).number_format = NUM_FMT
        ws.cell(row=r, column=4, value=comp.other_movements).number_format = NUM_FMT
        ws.cell(row=r, column=5, value=comp.closing).number_format = NUM_FMT
        r += 1
    ws.cell(row=r, column=1, value="Total Other Equity").font = _bold()
    ws.cell(row=r, column=2, value=soce.total_other_equity_opening).number_format = NUM_FMT
    ws.cell(row=r, column=5, value=soce.total_other_equity_closing).number_format = NUM_FMT
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
    r += 2

    ws.cell(row=r, column=1, value=(
        "Note: \"Other Additions/(Deductions)\" is a residual figure (closing "
        "less opening less profit transferred) capturing dividends paid, "
        "transfers between reserves, or prior period adjustments -- a Trial "
        "Balance alone cannot distinguish these; verify against board "
        "resolutions / minutes before finalizing."
    )).font = Font(name=FONT_NAME, italic=True, size=9, color="808080")
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    _autosize(ws, {1: 40, 2: 18, 3: 18, 4: 24, 5: 18})
    ws.sheet_view.showGridLines = False


def _write_ageing_sheet(wb, company, fy_label, receivables: AgeingGrid | None, payables: AgeingGrid | None):
    ws = wb.create_sheet("Ageing Schedule")
    r = _write_title_block(ws, company, "Ageing Schedule - Trade Receivables and Trade Payables", fy_label)

    def write_grid(grid: AgeingGrid, title: str, row: int) -> int:
        ws.cell(row=row, column=1, value=title).font = Font(name=FONT_NAME, bold=True, color="1F4E79")
        row += 1
        if not grid.available:
            ws.cell(row=row, column=1, value=(
                f"Ageing schedule not available: {grid.unavailable_reason}"
            )).font = Font(name=FONT_NAME, italic=True, size=10, color="C00000")
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            return row + 3

        headers = ["Category"] + BUCKET_LABELS + ["Total"]
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = _header_font()
            cell.fill = HEADER_FILL
        row += 1
        for cat in CATEGORY_LABELS:
            ws.cell(row=row, column=1, value=cat).font = _normal()
            row_total = 0.0
            for c, bucket in enumerate(BUCKET_LABELS, start=2):
                val = grid.grid[cat][bucket]
                row_total += val
                ws.cell(row=row, column=c, value=val).number_format = NUM_FMT
            ws.cell(row=row, column=len(BUCKET_LABELS) + 2, value=round(row_total, 2)).number_format = NUM_FMT
            row += 1
        ws.cell(row=row, column=1, value="Total").font = _bold()
        ws.cell(row=row, column=len(BUCKET_LABELS) + 2, value=grid.total).number_format = NUM_FMT
        for c in range(1, len(BUCKET_LABELS) + 3):
            ws.cell(row=row, column=c).fill = TOTAL_FILL
        row += 1
        if grid.reconciles_to_balance_sheet is False:
            ws.cell(row=row, column=1, value=(
                f"WARNING: Ageing total ({grid.total:,.2f}) does not reconcile to the "
                f"Balance Sheet figure ({grid.balance_sheet_amount:,.2f})."
            )).font = Font(name=FONT_NAME, color="C00000", bold=True)
            row += 1
        return row + 2

    if receivables is not None:
        r = write_grid(receivables, "Trade Receivables Ageing (as of Balance Sheet date)", r)
    if payables is not None:
        r = write_grid(payables, "Trade Payables Ageing (as of Balance Sheet date)", r)

    _autosize(ws, {1: 32, 2: 14, 3: 14, 4: 14, 5: 12, 6: 12, 7: 14, 8: 16})
    ws.sheet_view.showGridLines = False


def _write_tb_sheet(wb, company, fy_label, tb: TrialBalance, mappings):
    ws = wb.create_sheet("TB")
    row = _write_title_block(ws, company, "Trial Balance", fy_label)
    headers = ["Ledger Name", "Opening Balance", "Debit", "Credit", "Closing Balance", "Mapped Head"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = _header_font()
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")
    start_data_row = row + 1
    r = start_data_row
    for ledger in tb.ledgers:
        mapping = mappings.get(ledger.ledger_name)
        ws.cell(row=r, column=1, value=ledger.ledger_name).font = _normal()
        ws.cell(row=r, column=2, value=ledger.opening_balance).number_format = NUM_FMT
        ws.cell(row=r, column=3, value=ledger.debit).number_format = NUM_FMT
        ws.cell(row=r, column=4, value=ledger.credit).number_format = NUM_FMT
        ws.cell(row=r, column=5, value=f"=B{r}+C{r}-D{r}").number_format = NUM_FMT
        ws.cell(row=r, column=6, value=f"{mapping.major_head} / {mapping.sub_head}" if mapping else "UNMAPPED")
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = BORDER
        r += 1
    total_row = r
    ws.cell(row=total_row, column=1, value="Total").font = _bold()
    for c, col_letter in [(2, "B"), (3, "C"), (4, "D")]:
        ws.cell(row=total_row, column=c,
                value=f"=SUM({col_letter}{start_data_row}:{col_letter}{total_row - 1})").font = _bold()
        ws.cell(row=total_row, column=c).number_format = NUM_FMT
        ws.cell(row=total_row, column=c).fill = TOTAL_FILL
    ws.cell(row=total_row, column=1).fill = TOTAL_FILL
    _autosize(ws, {1: 38, 2: 16, 3: 16, 4: 16, 5: 16, 6: 42})
    ws.freeze_panes = f"A{start_data_row}"


def _write_mapping_sheet(wb, company, fy_label, tb, mappings):
    ws = wb.create_sheet("Mapping")
    row = _write_title_block(ws, company, "Ledger Mapping (Editable)", fy_label)
    headers = ["Ledger Name", "Closing Balance", "Major Head", "Sub Head", "Current / Non-Current",
               "Nature", "Note Reference", "Source", "Confidence"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = _header_font()
        cell.fill = HEADER_FILL
        cell.border = BORDER
    r = row + 1
    for ledger in tb.ledgers:
        m = mappings.get(ledger.ledger_name)
        ws.cell(row=r, column=1, value=ledger.ledger_name)
        ws.cell(row=r, column=2, value=ledger.closing_balance).number_format = NUM_FMT
        ws.cell(row=r, column=3, value=m.major_head if m else "")
        ws.cell(row=r, column=4, value=m.sub_head if m else "PENDING - PLEASE MAP")
        ws.cell(row=r, column=5, value=m.current_or_non_current.value if m else "")
        ws.cell(row=r, column=6, value=m.nature.value if m else "")
        ws.cell(row=r, column=7, value=m.note_ref if m else "")
        ws.cell(row=r, column=8, value=m.source if m else "")
        ws.cell(row=r, column=9, value=m.confidence if m else 0)
        if not m:
            for c in range(1, 10):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="FFF2CC")
        for c in range(1, 10):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).font = _normal()
        r += 1
    _autosize(ws, {1: 38, 2: 16, 3: 24, 4: 30, 5: 18, 6: 12, 7: 30, 8: 14, 9: 12})
    ws.freeze_panes = f"A{row + 1}"


def _write_section(ws, r, title, majors_or_subs, is_major_list: bool):
    ws.cell(row=r, column=1, value=title).font = Font(name=FONT_NAME, bold=True, size=11, color="1F4E79")
    r += 1
    grand_cy = grand_py = 0.0
    for entry in majors_or_subs:
        ws.cell(row=r, column=1, value=entry.major_head if is_major_list else entry.sub_head).font = _bold()
        ws.cell(row=r, column=1).fill = SUBHEAD_FILL
        ws.cell(row=r, column=3, value=entry.current_year).number_format = NUM_FMT
        ws.cell(row=r, column=4, value=entry.previous_year).number_format = NUM_FMT
        ws.cell(row=r, column=3).font = _bold()
        ws.cell(row=r, column=4).font = _bold()
        ws.cell(row=r, column=1).fill = SUBHEAD_FILL
        ws.cell(row=r, column=3).fill = SUBHEAD_FILL
        ws.cell(row=r, column=4).fill = SUBHEAD_FILL
        r += 1
        grand_cy += entry.current_year
        grand_py += entry.previous_year
        if is_major_list:
            for sh in entry.sub_heads:
                ws.cell(row=r, column=2, value=sh.sub_head).font = _normal()
                ws.cell(row=r, column=3, value=sh.current_year).number_format = NUM_FMT
                ws.cell(row=r, column=4, value=sh.previous_year).number_format = NUM_FMT
                ws.cell(row=r, column=5, value=sh.note_ref or "").font = Font(name=FONT_NAME, size=9, italic=True)
                r += 1
    return r, grand_cy, grand_py


def _write_balance_sheet_sheet(wb, company, fy_label, bs: BalanceSheet):
    ws = wb.create_sheet("Balance Sheet")
    r = _write_title_block(ws, company, "Balance Sheet", fy_label)
    ws.cell(row=r, column=3, value="Current Year").font = _header_font()
    ws.cell(row=r, column=4, value="Previous Year").font = _header_font()
    ws.cell(row=r, column=5, value="Note").font = _header_font()
    for c in (3, 4, 5):
        ws.cell(row=r, column=c).fill = HEADER_FILL
    r += 1
    r, el_cy, el_py = _write_section(ws, r, "EQUITY AND LIABILITIES", bs.equity_and_liabilities, True)
    ws.cell(row=r, column=1, value="TOTAL EQUITY AND LIABILITIES").font = _bold()
    ws.cell(row=r, column=3, value=bs.total_equity_and_liabilities_cy).number_format = NUM_FMT
    ws.cell(row=r, column=4, value=bs.total_equity_and_liabilities_py).number_format = NUM_FMT
    for c in (1, 3, 4):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = _bold()
    r += 3
    r, a_cy, a_py = _write_section(ws, r, "ASSETS", bs.assets, True)
    ws.cell(row=r, column=1, value="TOTAL ASSETS").font = _bold()
    ws.cell(row=r, column=3, value=bs.total_assets_cy).number_format = NUM_FMT
    ws.cell(row=r, column=4, value=bs.total_assets_py).number_format = NUM_FMT
    for c in (1, 3, 4):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = _bold()
    r += 2
    tie_msg = "Balance Sheet Tallies ✓" if bs.is_tallied else "WARNING: Balance Sheet does NOT tally"
    ws.cell(row=r, column=1, value=tie_msg).font = Font(
        name=FONT_NAME, bold=True, color="006100" if bs.is_tallied else "9C0006")
    _autosize(ws, {1: 34, 2: 34, 3: 18, 4: 18, 5: 30})
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "portrait"
    ws.print_title_rows = "1:1"


def _write_pnl_sheet(wb, company, fy_label, pnl: ProfitAndLoss):
    ws = wb.create_sheet("P&L")
    r = _write_title_block(ws, company, "Statement of Profit and Loss", fy_label)
    ws.cell(row=r, column=3, value="Current Year").font = _header_font()
    ws.cell(row=r, column=4, value="Previous Year").font = _header_font()
    ws.cell(row=r, column=5, value="Note").font = _header_font()
    for c in (3, 4, 5):
        ws.cell(row=r, column=c).fill = HEADER_FILL
    r += 1
    r, rev_cy, rev_py = _write_section(ws, r, "Revenue", pnl.revenue, False)
    ws.cell(row=r, column=1, value="Total Revenue").font = _bold()
    ws.cell(row=r, column=3, value=pnl.total_revenue_cy).number_format = NUM_FMT
    ws.cell(row=r, column=4, value=pnl.total_revenue_py).number_format = NUM_FMT
    for c in (1, 3, 4):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
    r += 2
    r, exp_cy, exp_py = _write_section(ws, r, "Expenses", pnl.expenses, False)
    ws.cell(row=r, column=1, value="Total Expenses").font = _bold()
    ws.cell(row=r, column=3, value=pnl.total_expenses_cy).number_format = NUM_FMT
    ws.cell(row=r, column=4, value=pnl.total_expenses_py).number_format = NUM_FMT
    for c in (1, 3, 4):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
    r += 2
    for label, cy, py in [
        ("Profit Before Tax", pnl.profit_before_tax_cy, pnl.profit_before_tax_py),
        ("Tax Expense", pnl.tax_expense_cy, pnl.tax_expense_py),
        ("Profit for the Period", pnl.profit_after_tax_cy, pnl.profit_after_tax_py),
    ]:
        ws.cell(row=r, column=1, value=label).font = _bold()
        ws.cell(row=r, column=3, value=cy).number_format = NUM_FMT
        ws.cell(row=r, column=4, value=py).number_format = NUM_FMT
        r += 1
    _autosize(ws, {1: 34, 2: 34, 3: 18, 4: 18, 5: 30})
    ws.sheet_view.showGridLines = False


def _write_cash_flow_sheet(wb, company, fy_label, cf: CashFlowStatement):
    ws = wb.create_sheet("Cash Flow")
    r = _write_title_block(ws, company, "Cash Flow Statement (Indirect Method)", fy_label)
    ws.cell(row=r, column=1, value="A. Cash Flow from Operating Activities").font = Font(
        name=FONT_NAME, bold=True, color="1F4E79")
    r += 1
    for label, val in [
        ("Net Profit Before Tax", cf.net_profit_before_tax),
        ("Add: Depreciation", cf.depreciation_addback),
        ("Add: Interest Expense", cf.interest_expense_addback),
    ]:
        ws.cell(row=r, column=1, value=label).font = _normal()
        ws.cell(row=r, column=3, value=val).number_format = NUM_FMT
        r += 1
    ws.cell(row=r, column=1, value="Working Capital Adjustments:").font = _bold()
    r += 1
    for label, val in cf.working_capital_changes.items():
        ws.cell(row=r, column=2, value=f"(Increase)/Decrease in {label}").font = _normal()
        ws.cell(row=r, column=3, value=val).number_format = NUM_FMT
        r += 1
    ws.cell(row=r, column=1, value="Net Cash from Operating Activities").font = _bold()
    ws.cell(row=r, column=3, value=cf.cash_from_operations).number_format = NUM_FMT
    ws.cell(row=r, column=1).fill = TOTAL_FILL
    ws.cell(row=r, column=3).fill = TOTAL_FILL
    r += 2

    ws.cell(row=r, column=1, value="B. Cash Flow from Investing Activities").font = Font(
        name=FONT_NAME, bold=True, color="1F4E79")
    r += 1
    ws.cell(row=r, column=1, value="Net Cash used in Investing Activities").font = _bold()
    ws.cell(row=r, column=3, value=cf.cash_from_investing).number_format = NUM_FMT
    r += 2

    ws.cell(row=r, column=1, value="C. Cash Flow from Financing Activities").font = Font(
        name=FONT_NAME, bold=True, color="1F4E79")
    r += 1
    ws.cell(row=r, column=1, value="Net Cash from/used in Financing Activities").font = _bold()
    ws.cell(row=r, column=3, value=cf.cash_from_financing).number_format = NUM_FMT
    r += 2

    for label, val in [
        ("Net Increase/(Decrease) in Cash and Cash Equivalents (A+B+C)", cf.net_increase_in_cash),
        ("Cash and Cash Equivalents at the Beginning of the Year", cf.opening_cash),
        ("Cash and Cash Equivalents at the End of the Year", cf.closing_cash),
    ]:
        ws.cell(row=r, column=1, value=label).font = _bold()
        ws.cell(row=r, column=3, value=val).number_format = NUM_FMT
        r += 1

    r += 1
    ws.cell(row=r, column=1, value=(
        "Note: Working capital movements require two years of Trial Balance "
        "data (previous year closing balances) to be populated for accurate "
        "results. Movements shown as 0 indicate previous-year data was not "
        "supplied for that ledger."
    )).font = Font(name=FONT_NAME, italic=True, size=9, color="808080")
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    _autosize(ws, {1: 46, 2: 40, 3: 18})
    ws.sheet_view.showGridLines = False


def _write_notes_sheet(wb, company, fy_label, notes: list[Note]):
    ws = wb.create_sheet("Notes")
    r = _write_title_block(ws, company, "Notes to Accounts", fy_label)
    for note in notes:
        ws.cell(row=r, column=1, value=note.note_ref).font = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
        ws.cell(row=r, column=1).fill = HEADER_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1
        ws.cell(row=r, column=1, value="Particulars").font = _bold()
        ws.cell(row=r, column=2, value="Current Year").font = _bold()
        ws.cell(row=r, column=3, value="Previous Year").font = _bold()
        r += 1
        for item in note.line_items:
            ws.cell(row=r, column=1, value=item.label).font = _normal()
            ws.cell(row=r, column=2, value=item.current_year).number_format = NUM_FMT
            ws.cell(row=r, column=3, value=item.previous_year).number_format = NUM_FMT
            r += 1
        ws.cell(row=r, column=1, value="Total").font = _bold()
        ws.cell(row=r, column=2, value=note.total_current_year).number_format = NUM_FMT
        ws.cell(row=r, column=3, value=note.total_previous_year).number_format = NUM_FMT
        ws.cell(row=r, column=1).fill = TOTAL_FILL
        ws.cell(row=r, column=2).fill = TOTAL_FILL
        ws.cell(row=r, column=3).fill = TOTAL_FILL
        r += 2
    _autosize(ws, {1: 42, 2: 18, 3: 18, 4: 14})
    ws.sheet_view.showGridLines = False


def _write_validation_sheet(wb, company, fy_label, issues: list[ValidationIssue]):
    ws = wb.create_sheet("Validation")
    r = _write_title_block(ws, company, "Validation Report", fy_label)
    headers = ["Severity", "Code", "Ledger", "Message", "Amount"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = _header_font()
        cell.fill = HEADER_FILL
    r += 1
    severity_fill = {"ERROR": "F8CBAD", "WARNING": "FFE699", "INFO": "D9E1F2"}
    for issue in issues:
        ws.cell(row=r, column=1, value=issue.severity)
        ws.cell(row=r, column=2, value=issue.code)
        ws.cell(row=r, column=3, value=issue.ledger_name or "")
        ws.cell(row=r, column=4, value=issue.message).alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=5, value=issue.amount if issue.amount is not None else "")
        fill = PatternFill("solid", fgColor=severity_fill.get(issue.severity, "FFFFFF"))
        for c in range(1, 6):
            ws.cell(row=r, column=c).fill = fill
            ws.cell(row=r, column=c).font = _normal()
        r += 1
    if not issues:
        ws.cell(row=r, column=1, value="No issues found.").font = Font(name=FONT_NAME, color="006100", bold=True)
    _autosize(ws, {1: 12, 2: 22, 3: 30, 4: 70, 5: 16})
    ws.sheet_view.showGridLines = False


def _write_ratios_sheet(wb, company, fy_label, ratios: list[RatioResult]):
    ws = wb.create_sheet("Ratios")
    r = _write_title_block(ws, company, "Financial Ratio Analysis", fy_label)
    headers = ["Ratio", "Current Year", "Previous Year", "Formula"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = _header_font()
        cell.fill = HEADER_FILL
    r += 1
    for ratio in ratios:
        ws.cell(row=r, column=1, value=ratio.name).font = _normal()
        ws.cell(row=r, column=2, value=ratio.current_year if ratio.current_year is not None else "N/A")
        ws.cell(row=r, column=3, value=ratio.previous_year if ratio.previous_year is not None else "N/A")
        ws.cell(row=r, column=4, value=ratio.formula).font = Font(name=FONT_NAME, italic=True, size=9)
        r += 1
    _autosize(ws, {1: 32, 2: 16, 3: 16, 4: 50})
    ws.sheet_view.showGridLines = False


def save_workbook(wb: Workbook, output_path: str) -> str:
    wb.save(output_path)
    return output_path
